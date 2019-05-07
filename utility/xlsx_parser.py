from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from utility.employees import Employee, Employees
from utility.resource_path import resource_path


class XLSXParser:
    """
    Класс XLSXParser служит для импорта XLSX файлов.
    """

    def __init__(self):
        self.__xlsx_filename = None
        self.__employees = Employees()
        self.__errors = list()
        self.__warnings = list()

    def __reset_state(self):
        self.__xlsx_filename = None
        self.__employees = Employees()
        self.__errors = list()
        self.__warnings = list()

    def get_employees(self):
        return self.__employees

    def get_errors(self):
        return self.__errors

    def get_warnings(self):
        return self.__warnings

    def load_file(self, filename):
        """Разбирает XLSX файл и заполняет данными список сотрудников"""
        self.__reset_state()

        # Пробуем открыть файл
        try:
            wb = load_workbook(filename)
            first_sheet_name = wb.sheetnames[0]
            ws = wb[first_sheet_name]
        except OSError:
            self.__errors.append('ERROR: XLSX файл по пути "' + filename + '" не найден.')
            return False
        self.__xlsx_filename = filename

        excel_employees = ws.rows

        for excel_employee in excel_employees:
            # Пропускаем шапку таблицы
            if excel_employee[0].row == 1:
                continue
            employee = self.__employee_from_row(excel_employee)
            self.__employees.add(employee)
        return True

    def __employee_from_row(self, excel_row):
        new_employee = Employee()
        employee_full_name = " ".join((str(excel_row[0].value), str(excel_row[1].value), str(excel_row[2].value)))
        for field in Employee.ALL_FIELDS:
            field_id = Employee.ALL_FIELDS.index(field)

            # Проверка форматов ячеек
            if str(type(excel_row[field_id].value)) != str(type(None)):
                if field == 'birth_date':
                    if str(type(excel_row[field_id].value)) != "<class 'datetime.datetime'>":
                        self.__warnings.append("WARNING: Формат ячейки 'Дата рождения' сотрудника {} не является датой."
                                               " Пожалуйста, проверьте формат ячейки. Она должа иметь формат 'Дата'."
                                               .format(employee_full_name))
                        excel_row[field_id].value = None
                elif field == 'experience':
                    if str(type(excel_row[field_id].value)) != "<class 'int'>":
                        self.__warnings.append("WARNING: Формат ячейки 'Стаж' сотрудника {} не является числом. "
                                               "Пожалуйста, проверьте формат ячейки. Она должа иметь формат 'Числовой'."
                                               .format(employee_full_name))
                        excel_row[field_id].value = ''
                else:
                    if str(type(excel_row[field_id].value)) != "<class 'str'>":
                        self.__warnings.append("WARNING: Формат ячейки '{}' сотрудника {} не является текстовым. "
                                               "Пожалуйста, проверьте формат ячейки. Она должа иметь формат 'Текст'."
                                               .format(Employee.translate(field), employee_full_name))
                        excel_row[field_id].value = ''

            # аие полей сотрудника
            if field in Employee.LIST_FIELDS:
                if excel_row[field_id].value is None:
                    value = list()
                else:
                    value = str(excel_row[field_id].value).split(',')
                    value = list(map(str.strip, value))
            else:
                value = '' if excel_row[field_id].value is None else excel_row[field_id].value

            if field in Employee.LIST_FIELDS:
                new_employee[field] = value
            elif field == 'birth_date':
                if value != '':
                    value = value.strftime("%Y-%m-%d")
                new_employee[field] = value
            elif field == 'sex':
                if value in ('Мужской', 'Женский', ''):
                    new_employee[field] = value
                else:
                    self.__warnings.append("WARNING: Ячейка 'Пол' сотрудника {} заполнена неправильно. "
                                           "Пожалуйста, проверьте текст ячейки. "
                                           "Она должа содержать одно из двух значений: 'Мужской' или 'Женский'."
                                           .format(employee_full_name))
                    new_employee[field] = ''
            else:
                new_employee[field] = str(value)
        return new_employee

    @staticmethod
    def save_to_file(filename, employees):
        """Сохраняет данные в XLSX файл"""
        try:
            wb = load_workbook(resource_path('./resources/template_excel.xlsx'))
            first_sheet_name = wb.sheetnames[0]
            ws = wb[first_sheet_name]
        except OSError as e:
            print(e)
            return False

        # Заполняем данные сотрудников
        ws._current_row = 1  # Фикс бага в openpyxl, который не правильно считает текущую строку
        for employee in employees.values():
            xlsx_employee = XLSXParser.__employee_to_row(employee)
            ws.append(xlsx_employee)
        for row in ws.iter_rows(min_row=2):
            XLSXParser.__apply_styles_to_row(row)
            rd = ws.row_dimensions[row[0].row]
            rd.height = 40
        wb.save(filename=filename)
        return True

    @staticmethod
    def __employee_to_row(employee):
        xlsx_employee = list()
        for field in Employee.ALL_FIELDS:
            if field == 'birth_date':
                birth_date = employee['birth_date']
                birth_date = datetime.strptime(birth_date, "%Y-%m-%d")
                # birth_date = birth_date.strftime("%d.%m.%Y")
                xlsx_employee.append(birth_date)
            elif field == 'experience':
                experience = '' if employee['experience'] == '' else int(employee['experience'])
                xlsx_employee.append(experience)
            elif field in Employee.LIST_FIELDS:
                hazards = employee[field]
                hazards_str = ', '.join(hazards)
                xlsx_employee.append(hazards_str)
            else:
                value = employee[field]
                xlsx_employee.append(value)
        return xlsx_employee

    @staticmethod
    def __apply_styles_to_row(row):
        border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))

        for cell in row:
            cell.border = border
            if cell.column in (4, 5, 7, 8):
                alignment = Alignment(horizontal='center', vertical='center')
                cell.alignment = alignment
            else:
                alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.alignment = alignment

            if cell.column == 5:
                cell.number_format = 'DD.MM.YYYY'
            elif cell.column == 7:
                cell.number_format = '0'
